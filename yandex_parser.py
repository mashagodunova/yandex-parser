import requests
import os
import openpyxl
from io import BytesIO
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
token = '5643122315:AAHTc3MOGgFiwXznLctdNPVRZ6QDO_iDSSs'


def create_driver():
    options = Options()
    options.add_argument(
        f"user-data-dir=Profile_14")
    options.add_argument('--user-agent="Mozilla/5.0 (Windows Phone 10.0; Android 4.2.1; Microsoft; Lumia 640 XL LTE) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Mobile Safari/537.36 Edge/12.10166"')
    #options.add_argument("--headless")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--v=99")
    options.add_argument("--no-sandbox")
    driver = uc.Chrome(options=options)
    return driver


def feedback_yandex_market(link):
    driver = create_driver()
    driver.get(link)
    time.sleep(10)
    feedback_arr = []
    rating_feedback = {
        'Отличный товар': 5,
        'Хороший товар': 4,
        'Обычный товар': 3,
        'Плохой товар': 2,
        'Ужасный товар': 1
    }

    try:
        feedbacks = driver.find_elements(By.XPATH,"//div[@class='_13uSY']")
        for feedback in feedbacks:
            feedback_dict = {}
            try:
                image_author = feedback.find_element_by_class_name(
                    '_3CWdE').get_attribute('src')
                image_author = original_photo_size(image_author)
            except:
                image_author = 'Нет аватарки'
            feedback_dict['avatar'] = image_author
            name_user = feedback.find_element_by_class_name('_1mJcZ')
            feedback_dict['name'] = name_user.text
            try:
                image_feedback = feedback.find_element_by_class_name(
                    '_1mPav').get_attribute('src')
                image_feedback  = original_photo_size(image_feedback)
            except:
                image_feedback = 'Нет фото у отзыва'
            feedback_dict['photo_feed'] = image_feedback
            #text_feedback = feedback.find_element_by_class_name('_3IXcz')
            try:
                text_feedback = feedback.find_element_by_class_name(
                    '_3IXcz').text
            except:
                text_feedback = 'Нет текста у отзыва'
            feedback_dict['text_feedback'] = text_feedback
            data_feedback = feedback.find_element_by_class_name('kx7am')
            feedback_dict['data_feedback'] = data_feedback.text
            rating = feedback.find_element_by_class_name('pcIgr')
            feedback_dict['rating'] = rating_feedback[rating.text]
            feedback_arr.append(feedback_dict)
        driver.quit()
        return feedback_arr
    except Exception as e:
        print(f'Ошибка парсинга {e}')
        driver.quit()
        return f'error'


def parser_yandex_market(link):
    driver = create_driver()
    driver.get(link)
    time.sleep(10)
    dict_market = {}
    try:
        print(
            'Название, цена, рейтинг, характеристики рядом с заголовком, коротко о товаре')
        
        name = driver.find_element(By.XPATH,"//h1[@class='_3TVFy _2SUA6 jM85b _13aK2 _1A5yJ']").text
        try:
            rating = driver.find_element(By.XPATH,"//span[@class='_24dJu']").text
        except:
            rating = 'Нет данных'
       
        feed_bck_btn = driver.find_element(By.XPATH,"//a[@class=''EQlfk COmSo'']")
        link_feed_bck_btn = feed_bck_btn.get_attribute('href')
        feed_bck_btn.click()
        time.sleep(10)

        feedback_page = driver.find_element(By.XPATH,"//a[@class=''EQlfk COmSo'']")
        if len(feedback_page) == 0:
            print(link_feed_bck_btn)
            dict_market['feedback'] = [link_feed_bck_btn]
        else:
            feedbck_link = []
            for i in range(len(feedback_page)):
                try:
                    print(feedback_page[i].get_attribute('href'))
                    if feedback_page[i].get_attribute('href') not in feedbck_link:
                        feedbck_link.append(feedback_page[i].get_attribute('href'))
                except:
                    pass
            dict_market['feedback'] = feedbck_link
        driver.quit()
        return dict_market
    except Exception as e:
        print(f'Ошибка парсинга {e}')
        driver.quit()
        return 'error',e


def start_parser(link):
    print('We started')
    quetstion_func,e = parser_yandex_market(link)
    print('We started')
    if quetstion_func == 'error':
        print('Ошибка при парсинге, повторите попытку',e)
    else:
        wb = openpyxl.Workbook()
        # Основная информация
        main_page = wb.create_sheet(
            title='Основная информация', index=0)
        sheet_main_page = wb['Основная информация']
        field_name_main_page = ['Название товара', 'Цена', 'Рейтинг',
                                'Выбор покупателей', 'Характеристики, которые пользователи отметили в отзывах', 'Краткая характеристика', 'Характеристика', 'Отзыв Яндекса']

        photo_field_main_page = ['№','Фото']
        sheet_main_page.append(field_name_main_page)

        sheet_main_page.append([quetstion_func['name'], quetstion_func['price'], quetstion_func['rating'],
                                quetstion_func['user_change'], quetstion_func['characteristic_h1'], quetstion_func['characteristic_low'], quetstion_func['characteristic_product'], quetstion_func['text_feedback']])
        sheet_main_page.append(['', '', '', '', '','', '',''])
        sheet_main_page.append(photo_field_main_page)
        photo_count = 0
        for image in quetstion_func['images']:
            photo_count +=1 
            sheet_main_page.append([photo_count, image])
        # Страница вопросов
        feedback_page = wb.create_sheet(
            title='Отзывы', index=0)
        sheet_feedbck_page = wb['Отзывы']
        field_name_feedbck_page = ['Авaтарка', 'Имя пользователя', 'Фото отзыва', 'Текст',
                                   'Дата', 'Рейтинг']
        sheet_feedbck_page.append(field_name_feedbck_page)
        for i in quetstion_func['feedback']:
            if i != 'error':
                arr_feedback = feedback_yandex_market(i)
                for index in arr_feedback:
                    sheet_feedbck_page.append([index['avatar'], index['name'], index['photo_feed'],
                                               index['text_feedback'], index['data_feedback'], index['rating']])
        filename = quetstion_func['name']
        excelfile = f'{filename}.xlsx'
        if '/' in excelfile:
            excelfile = excelfile.replace('/', ' ')
        wb.save(excelfile)
        # status = requests.post(
        #     f'https://api.telegram.org/bot{token}/sendDocument?chat_id={chat_id}', files={'document': open(excelfile, 'rb')})
        # print(status)
        # os.remove(excelfile)
start_parser('https://market.yandex.ru/product--nabor-iz-10-par-noskov-moscowsocksclub-m20-miks-iskusstvo-razmer-27-41-43/1439308463?glfilter=14474426%3AMjcgKDQxLTQzKQ_101447518663&glfilter=14871214%3A15098577_101447518663&cpc=zHewvSNgIjPtJC9iD60JL3m43OcnhUqQ-g6C4kDb05rWnvIKBybbR6unOSuLlBsL9g15fnalFtreycesyu-TgSxkTsBYWCoNIxS5QvFMzLDWg2VG_56D7oYSPcsjO6rlKiAnV7kSCqmriKKOOtdM0Vd8TVPSkApivEnqem6eGihq0VwesfusPWNxg1F5Tq3i&sku=101447518663&offerid=7JFFcigWlvMzkapCVGazig&cpa=1')
